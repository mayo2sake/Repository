import argparse
import json
import math
import re
import time
from dataclasses import dataclass
from datetime import datetime, timezone
from html import escape
from pathlib import Path
from typing import Dict, List, Tuple
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_URL = "https://www.e-typing.ne.jp/ranking/"
LIST_URL = urljoin(BASE_URL, "ranking_list.asp?im=0")
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0 Safari/537.36"
)


@dataclass
class RankingInfo:
    period_key: str
    label: str
    list_title: str
    no: str
    url: str


def fetch_text(session: requests.Session, url: str, *, method: str = "get", data=None) -> str:
    response = session.post(url, data=data, timeout=30) if method == "post" else session.get(url, timeout=30)
    response.raise_for_status()
    response.encoding = response.apparent_encoding or "utf-8"
    return response.text


def parse_ranking_list(html: str) -> List[RankingInfo]:
    soup = BeautifulSoup(html, "html.parser")
    rankings: List[RankingInfo] = []
    for index, link in enumerate(soup.select("ul.ranking a[href]")):
        href = link.get("href", "")
        no_match = re.search(r"[?&]no=(\d+)", href)
        if not no_match:
            continue
        period_key = "last_week" if index == 0 else "week_before" if index == 1 else f"history_{index + 1}"
        label = "先週" if index == 0 else "先々週" if index == 1 else f"過去{index + 1}"
        rankings.append(
            RankingInfo(
                period_key=period_key,
                label=label,
                list_title=link.get_text(strip=True),
                no=no_match.group(1),
                url=urljoin(BASE_URL, href),
            )
        )
    return rankings


def parse_page(html: str) -> Tuple[str, int, List[dict]]:
    soup = BeautifulSoup(html, "html.parser")
    section = soup.select_one("section#ranking")
    title = section.select_one("h1").get_text(" ", strip=True) if section else ""

    max_match = re.search(r"ps_page_max\s*=\s*(\d+)", html)
    page_max = int(max_match.group(1)) if max_match else 1

    rows = []
    for item in soup.select("ul.ranking > li"):
        if "head" in item.get("class", []):
            continue
        rank = item.select_one(".rank")
        user = item.select_one(".user")
        score = item.select_one(".score")
        if not (rank and user and score):
            continue
        rank_text = rank.get_text(strip=True)
        rows.append(
            {
                "rank": int(re.sub(r"\D", "", rank_text) or "0"),
                "rank_text": rank_text,
                "user": user.get_text(" ", strip=True),
                "score": int(re.sub(r"\D", "", score.get_text(strip=True)) or "0"),
            }
        )
    return title, page_max, rows


def scrape_ranking(info: RankingInfo, wait_seconds: float) -> Tuple[dict, List[dict]]:
    session = requests.Session()
    session.headers.update({"User-Agent": USER_AGENT})

    first_html = fetch_text(session, info.url)
    title, page_max, all_rows = parse_page(first_html)

    post_url = urljoin(BASE_URL, "trysc.asp")
    for page in range(2, page_max + 1):
        html = fetch_text(session, post_url, method="post", data={"f_pg": str(page), "f_pg_sz": ""})
        _, _, rows = parse_page(html)
        all_rows.extend(rows)
        if wait_seconds:
            time.sleep(wait_seconds)

    summary = {
        "label": info.label,
        "title": title,
        "no": info.no,
        "source_url": info.url,
        "page_count": page_max,
        "row_count": len(all_rows),
        "generated_at": datetime.now(timezone.utc).isoformat(),
    }
    return summary, all_rows


def population_stddev(values: List[int], average: float) -> float:
    if not values:
        return 0.0
    variance = sum((value - average) ** 2 for value in values) / len(values)
    return math.sqrt(variance)


def build_stats(rows: List[dict]) -> Dict[str, float]:
    scores = [row["score"] for row in rows]
    participant_count = len(scores)
    average = sum(scores) / participant_count if participant_count else 0.0
    stddev = population_stddev(scores, average)
    return {
        "participant_count": participant_count,
        "average": average,
        "stddev": stddev,
        "hensachi_60": average + stddev,
        "hensachi_70": average + stddev * 2,
        "hensachi_80": average + stddev * 3,
        "hensachi_90": average + stddev * 4,
        "max_score": max(scores) if scores else 0,
        "min_score": min(scores) if scores else 0,
    }


def build_distribution(rows: List[dict], width: int = 50) -> List[dict]:
    scores = [row["score"] for row in rows]
    if not scores:
        return []
    start = (min(scores) // width) * width
    end = ((max(scores) // width) + 1) * width
    buckets = []
    for bucket_start in range(start, end, width):
        bucket_end = bucket_start + width - 1
        count = sum(1 for score in scores if bucket_start <= score <= bucket_end)
        buckets.append({"label": f"{bucket_start}-{bucket_end}", "count": count})
    return buckets


def write_workbook(output_path: Path, summary: dict, rows: List[dict]) -> None:
    stats = build_stats(rows)
    workbook = Workbook()

    summary_sheet = workbook.active
    summary_sheet.title = "概要"
    summary_sheet["A1"] = summary["title"]
    summary_sheet["A1"].font = Font(bold=True, size=14)
    summary_items = [
        ("参加人数", stats["participant_count"]),
        ("平均", round(stats["average"], 2)),
        ("標準偏差", round(stats["stddev"], 2)),
        ("", ""),
        ("偏差値60", round(stats["hensachi_60"], 2)),
        ("偏差値70", round(stats["hensachi_70"], 2)),
        ("偏差値80", round(stats["hensachi_80"], 2)),
        ("偏差値90", round(stats["hensachi_90"], 2)),
        ("", ""),
        ("取得元URL", summary["source_url"]),
        ("ページ数", summary["page_count"]),
    ]
    for row_index, (label, value) in enumerate(summary_items, start=3):
        summary_sheet.cell(row=row_index, column=1, value=label)
        summary_sheet.cell(row=row_index, column=2, value=value)
    summary_sheet.column_dimensions["A"].width = 18
    summary_sheet.column_dimensions["B"].width = 72

    detail_sheet = workbook.create_sheet("詳細結果")
    detail_sheet.append(["順位", "順位表示", "ユーザー名", "スコア"])
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in detail_sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for row in rows:
        detail_sheet.append([row["rank"], row["rank_text"], str(row["user"]), row["score"]])
        detail_sheet.cell(row=detail_sheet.max_row, column=3).number_format = "@"
    detail_sheet.freeze_panes = "A2"
    detail_sheet.auto_filter.ref = detail_sheet.dimensions
    for column_index, width in enumerate([10, 12, 42, 10], start=1):
        detail_sheet.column_dimensions[get_column_letter(column_index)].width = width

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def read_manifest(path: Path) -> List[dict]:
    if not path.exists():
        return []
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, data) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def upsert_manifest(manifest: List[dict], item: dict) -> List[dict]:
    by_no = {entry["no"]: entry for entry in manifest}
    by_no[item["no"]] = item
    return sorted(by_no.values(), key=lambda entry: int(entry["no"]), reverse=True)


def render_page(title: str, body: str, relative_root: str = ".") -> str:
    return f"""<!doctype html>
<html lang="ja">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{escape(title)}</title>
  <link rel="stylesheet" href="{relative_root}/assets/style.css">
</head>
<body>
  <header class="site-header">
    <a class="brand" href="{relative_root}/index.html">e-typing 腕試し統計</a>
  </header>
  <main>
    {body}
  </main>
</body>
</html>
"""


def render_bar_svg(labels: List[str], values: List[float], color: str = "#2f80ed") -> str:
    width, height = 860, 360
    left, right, top, bottom = 52, 24, 20, 72
    chart_width = width - left - right
    chart_height = height - top - bottom
    max_value = max(values) if values else 1
    step = chart_width / max(len(values), 1)
    bar_width = max(6, step * 0.68)
    bars = []
    for index, value in enumerate(values):
        bar_height = 0 if max_value == 0 else chart_height * value / max_value
        x = left + index * step + (step - bar_width) / 2
        y = top + chart_height - bar_height
        label = escape(labels[index])
        bars.append(
            f'<rect x="{x:.2f}" y="{y:.2f}" width="{bar_width:.2f}" height="{bar_height:.2f}" fill="{color}">'
            f"<title>{label}: {value}</title></rect>"
        )
        if len(values) <= 24:
            bars.append(
                f'<text x="{x + bar_width / 2:.2f}" y="{height - 34}" text-anchor="end" '
                f'transform="rotate(-45 {x + bar_width / 2:.2f} {height - 34})">{label}</text>'
            )
    return (
        f'<svg class="chart-svg" viewBox="0 0 {width} {height}" role="img">'
        f'<line x1="{left}" y1="{top + chart_height}" x2="{width - right}" y2="{top + chart_height}" class="axis"/>'
        f'<line x1="{left}" y1="{top}" x2="{left}" y2="{top + chart_height}" class="axis"/>'
        f'<text x="{left - 8}" y="{top + 12}" text-anchor="end">{int(max_value)}</text>'
        + "".join(bars)
        + "</svg>"
    )


def render_horizontal_bar_svg(labels: List[str], values: List[float], color: str = "#16a34a") -> str:
    width = 860
    row_height = 24
    left, right, top, bottom = 92, 28, 18, 18
    height = top + bottom + row_height * max(len(values), 1)
    chart_width = width - left - right
    max_value = max(values) if values else 1
    rows = []
    for index, value in enumerate(values):
        y = top + index * row_height
        bar_width = 0 if max_value == 0 else chart_width * value / max_value
        label = escape(labels[index])
        rows.append(f'<text x="{left - 8}" y="{y + 15}" text-anchor="end">{label}</text>')
        rows.append(f'<rect x="{left}" y="{y + 3}" width="{bar_width:.2f}" height="16" fill="{color}"><title>{label}: {value}</title></rect>')
        rows.append(f'<text x="{left + bar_width + 6:.2f}" y="{y + 15}">{int(value)}</text>')
    return f'<svg class="chart-svg" viewBox="0 0 {width} {height}" role="img">{"".join(rows)}</svg>'


def render_line_svg(manifest: List[dict]) -> str:
    width, height = 860, 360
    left, right, top, bottom = 52, 28, 28, 56
    chart_width = width - left - right
    chart_height = height - top - bottom
    ordered = list(reversed(manifest))
    if not ordered:
        return '<div class="empty">データがありません</div>'
    values_a = [item["stats"]["average"] for item in ordered]
    values_b = [item["stats"]["hensachi_70"] for item in ordered]
    all_values = values_a + values_b
    min_value = min(all_values)
    max_value = max(all_values)
    if min_value == max_value:
        min_value -= 1
        max_value += 1

    def points(values: List[float]) -> str:
        coords = []
        denom = max(len(values) - 1, 1)
        for index, value in enumerate(values):
            x = left + chart_width * index / denom
            y = top + chart_height - chart_height * (value - min_value) / (max_value - min_value)
            coords.append(f"{x:.2f},{y:.2f}")
        return " ".join(coords)

    labels = []
    for index, item in enumerate(ordered):
        if len(ordered) <= 12 or index in (0, len(ordered) - 1):
            denom = max(len(ordered) - 1, 1)
            x = left + chart_width * index / denom
            labels.append(f'<text x="{x:.2f}" y="{height - 22}" text-anchor="middle">第{escape(item["no"])}回</text>')
    return f"""
<svg class="chart-svg" viewBox="0 0 {width} {height}" role="img">
  <line x1="{left}" y1="{top + chart_height}" x2="{width - right}" y2="{top + chart_height}" class="axis"/>
  <line x1="{left}" y1="{top}" x2="{left}" y2="{top + chart_height}" class="axis"/>
  <text x="{left - 8}" y="{top + 12}" text-anchor="end">{max_value:.0f}</text>
  <text x="{left - 8}" y="{top + chart_height}" text-anchor="end">{min_value:.0f}</text>
  <polyline points="{points(values_a)}" fill="none" stroke="#2f80ed" stroke-width="3"/>
  <polyline points="{points(values_b)}" fill="none" stroke="#dc2626" stroke-width="3"/>
  <circle cx="{left}" cy="14" r="5" fill="#2f80ed"/><text x="{left + 10}" y="18">平均</text>
  <circle cx="{left + 72}" cy="14" r="5" fill="#dc2626"/><text x="{left + 82}" y="18">偏差値70</text>
  {''.join(labels)}
</svg>
"""


def stat_cards(stats: dict) -> str:
    items = [
        ("参加人数", f"{stats['participant_count']:,}"),
        ("平均", f"{stats['average']:.2f}"),
        ("標準偏差", f"{stats['stddev']:.2f}"),
        ("最高スコア", f"{stats['max_score']:,}"),
        ("偏差値60", f"{stats['hensachi_60']:.2f}"),
        ("偏差値70", f"{stats['hensachi_70']:.2f}"),
        ("偏差値80", f"{stats['hensachi_80']:.2f}"),
        ("偏差値90", f"{stats['hensachi_90']:.2f}"),
    ]
    return "\n".join(f"<div class=\"stat\"><span>{label}</span><strong>{value}</strong></div>" for label, value in items)


def render_ranking_html(summary: dict, rows: List[dict], excel_name: str) -> str:
    stats = build_stats(rows)
    distribution = build_distribution(rows)
    top_rows = rows[:20]
    distribution_svg = render_bar_svg([item["label"] for item in distribution], [item["count"] for item in distribution])
    top_svg = render_horizontal_bar_svg([row["rank_text"] for row in top_rows], [row["score"] for row in top_rows])
    body = f"""
<section class="hero">
  <p class="eyebrow">第{escape(summary['no'])}回</p>
  <h1>{escape(summary['title'])}</h1>
  <div class="actions"><a class="button" href="../files/{escape(excel_name)}">Excelをダウンロード</a></div>
</section>
<section class="stats-grid">{stat_cards(stats)}</section>
<section class="chart-grid">
  <article class="panel"><h2>スコア分布</h2>{distribution_svg}</article>
  <article class="panel"><h2>上位20件</h2>{top_svg}</article>
</section>
<section class="panel">
  <h2>上位一覧</h2>
  <table>
    <thead><tr><th>順位</th><th>ユーザー名</th><th>スコア</th></tr></thead>
    <tbody>
      {''.join(f"<tr><td>{row['rank_text']}</td><td>{escape(str(row['user']))}</td><td>{row['score']}</td></tr>" for row in top_rows)}
    </tbody>
  </table>
</section>
"""
    return render_page(summary["title"], body, "..")


def render_index_html(manifest: List[dict]) -> str:
    latest = manifest[0] if manifest else None
    latest_block = ""
    if latest:
        latest_block = f"""
<section class="hero">
  <p class="eyebrow">最新公開</p>
  <h1>{escape(latest['title'])}</h1>
  <div class="actions"><a class="button" href="rankings/{escape(latest['no'])}.html">詳細を見る</a></div>
</section>
<section class="stats-grid">{stat_cards(latest['stats'])}</section>
"""
    rows = "".join(
        f"<tr><td>第{escape(item['no'])}回</td><td><a href=\"rankings/{escape(item['no'])}.html\">{escape(item['title'])}</a></td><td>{item['stats']['average']:.2f}</td><td>{item['stats']['stddev']:.2f}</td></tr>"
        for item in manifest
    )
    body = f"""
{latest_block}
<section class="panel">
  <h2>平均スコアの推移</h2>
  {render_line_svg(manifest)}
</section>
<section class="panel">
  <h2>公開履歴</h2>
  <table>
    <thead><tr><th>回</th><th>ランキング</th><th>平均</th><th>標準偏差</th></tr></thead>
    <tbody>{rows}</tbody>
  </table>
</section>
"""
    return render_page("e-typing 腕試し統計", body, ".")


def write_site_assets(docs_dir: Path) -> None:
    css = """
:root {
  color-scheme: light;
  --bg: #f6f8fb;
  --panel: #ffffff;
  --text: #172033;
  --muted: #5c677d;
  --line: #dfe5ee;
  --blue: #2f80ed;
}
* { box-sizing: border-box; }
body {
  margin: 0;
  background: var(--bg);
  color: var(--text);
  font-family: system-ui, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
}
.site-header {
  height: 56px;
  display: flex;
  align-items: center;
  padding: 0 24px;
  background: #fff;
  border-bottom: 1px solid var(--line);
}
.brand { color: var(--text); font-weight: 700; text-decoration: none; }
main { width: min(1120px, calc(100% - 32px)); margin: 28px auto 56px; }
.hero { margin-bottom: 20px; }
.eyebrow { margin: 0 0 6px; color: var(--blue); font-weight: 700; }
h1 { margin: 0; font-size: clamp(24px, 4vw, 40px); letter-spacing: 0; }
h2 { margin: 0 0 16px; font-size: 18px; }
.actions { margin-top: 16px; }
.button {
  display: inline-flex;
  align-items: center;
  min-height: 40px;
  padding: 0 14px;
  background: var(--blue);
  color: white;
  text-decoration: none;
  border-radius: 6px;
  font-weight: 700;
}
.stats-grid {
  display: grid;
  grid-template-columns: repeat(auto-fit, minmax(150px, 1fr));
  gap: 10px;
  margin: 20px 0;
}
.stat, .panel {
  background: var(--panel);
  border: 1px solid var(--line);
  border-radius: 8px;
}
.stat { padding: 14px; }
.stat span { display: block; color: var(--muted); font-size: 13px; }
.stat strong { display: block; margin-top: 6px; font-size: 24px; }
.chart-grid {
  display: grid;
  grid-template-columns: minmax(0, 1fr) minmax(0, 1fr);
  gap: 16px;
  margin-bottom: 16px;
}
.panel { padding: 18px; margin-bottom: 16px; }
.chart-svg { display: block; width: 100%; height: auto; overflow: visible; }
.chart-svg text { fill: var(--muted); font-size: 12px; }
.axis { stroke: var(--line); stroke-width: 1; }
table { width: 100%; border-collapse: collapse; }
th, td { padding: 10px 8px; border-bottom: 1px solid var(--line); text-align: left; }
th { color: var(--muted); font-size: 13px; }
@media (max-width: 760px) {
  .chart-grid { grid-template-columns: 1fr; }
  .site-header { padding: 0 16px; }
}
"""
    path = docs_dir / "assets" / "style.css"
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(css, encoding="utf-8")


def choose_period(period: str, rankings: List[RankingInfo]) -> RankingInfo:
    choices = {ranking.period_key: ranking for ranking in rankings[:2]}
    return choices[period]


def build(period: str, docs_dir: Path, wait: float) -> None:
    session = requests.Session()
    session.headers.update({"User-Agent": USER_AGENT})
    rankings = parse_ranking_list(fetch_text(session, LIST_URL))
    if len(rankings) < 2:
        raise RuntimeError("ランキング一覧から先週・先々週を特定できませんでした。")

    target = choose_period(period, rankings)
    summary, rows = scrape_ranking(target, wait)
    stats = build_stats(rows)
    excel_name = f"e_typing_trysc_{summary['no']}.xlsx"

    write_workbook(docs_dir / "files" / excel_name, summary, rows)
    write_json(docs_dir / "data" / f"{summary['no']}.json", {"summary": summary, "stats": stats, "rows": rows})
    write_site_assets(docs_dir)

    manifest_path = docs_dir / "data" / "rankings.json"
    manifest = read_manifest(manifest_path)
    manifest_item = {
        "no": summary["no"],
        "title": summary["title"],
        "source_url": summary["source_url"],
        "row_count": summary["row_count"],
        "page_count": summary["page_count"],
        "generated_at": summary["generated_at"],
        "excel": f"files/{excel_name}",
        "stats": stats,
    }
    manifest = upsert_manifest(manifest, manifest_item)
    write_json(manifest_path, manifest)

    ranking_html = render_ranking_html(summary, rows, excel_name)
    ranking_path = docs_dir / "rankings" / f"{summary['no']}.html"
    ranking_path.parent.mkdir(parents=True, exist_ok=True)
    ranking_path.write_text(ranking_html, encoding="utf-8")
    (docs_dir / "index.html").write_text(render_index_html(manifest), encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description="Build the e-typing ranking GitHub Pages site.")
    parser.add_argument("--period", choices=["last_week", "week_before"], default="last_week")
    parser.add_argument("--docs-dir", default="docs")
    parser.add_argument("--wait", type=float, default=0.1)
    args = parser.parse_args()
    build(args.period, Path(args.docs_dir), args.wait)


if __name__ == "__main__":
    main()
